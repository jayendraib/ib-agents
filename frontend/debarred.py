import pandas as pd
import os
import subprocess
import tempfile
import zipfile
import shutil
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import os
import subprocess
import tempfile
import zipfile
import shutil
import base64
import requests
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# ========== API CONFIGURATION ==========
QA_API_URL   = "https://qa-kychecker.indiabonds.com/SEBI/FileUploadApi"
PROD_API_URL = "https://kyc-kychecker.indiabonds.com/SEBI/FileUploadApi"
API_URL = QA_API_URL  # default fallback

QA_API_KEY   = "0b80d0fe42644ae3a5a8277ccff4254b"
PROD_API_KEY = "2ff65e9edd17473499e9c8f32a304190"
API_KEY = QA_API_KEY  # default fallback

COMMENT = "AutomatedUpload"

# ========== CONFIGURATION ==========
# 🔴 CHANGED: PAN No → PAN No.
MASTER_COLUMNS = ['PAN No.', 'Remark', 'Notice']

COLUMN_KEYWORDS = {
    'PAN No.': ['pan', 'pan no', 'pan number', 'panno', 'permanent account number'],
    'Remark': ['remark', 'remarks', 'order particular', 'Order Particulars','order particulars' 'description', 'particulars', 'particular','Subject','subject'],
    'Notice': ['notice', 'period', 'notice period', 'assessment year', 'ay', 'financial year', 'fy', 'notice no','status']
}

# ========== ZIP EXTRACTION ==========
def extract_zip(zip_path: str) -> tuple[list[str], str]:
    """
    Extracts ZIP and returns (excel_file_paths_list, temp_dir_path).
    Returns STRINGS only, not Path objects.
    """
    zip_path = Path(zip_path)
    print(f"\n📦 Found ZIP: {zip_path.name}")
    temp_dir = None

    # Method 1: System unzip
    try:
        result = subprocess.run(
            ["unzip", "-t", str(zip_path)],
            capture_output=True,
            text=True,
            timeout=10
        )
        if result.returncode == 0:
            temp_dir = Path(tempfile.mkdtemp(prefix="excel_zip_"))
            subprocess.run(
                ["unzip", "-o", str(zip_path), "-d", str(temp_dir)],
                check=True,
                capture_output=True
            )
            
            excel_files = []
            for file_path in temp_dir.rglob("*"):
                if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls']:
                    excel_files.append(str(file_path.resolve()))
            
            print(f"   📊 Extracted {len(excel_files)} Excel file(s) from ZIP")
            return excel_files, str(temp_dir)
    except Exception:
        pass

    # Method 2: Python fallback
    try:
        temp_dir = Path(tempfile.mkdtemp(prefix="excel_zip_py_"))
        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(temp_dir)
        
        excel_files = []
        for file_path in temp_dir.rglob("*"):
            if file_path.is_file() and file_path.suffix.lower() in ['.xlsx', '.xls']:
                excel_files.append(str(file_path.resolve()))
        
        print(f"   📊 Extracted {len(excel_files)} Excel file(s) from ZIP (fallback)")
        return excel_files, str(temp_dir)
        
    except Exception as e:
        print(f"   ❌ Failed to extract ZIP: {e}")
        return [], None

# ========== GET ALL EXCEL FILES ==========
def get_excel_files(paths):
    """
    Accepts files or folders.
    - Folder: scans for Excel files AND ZIP files inside
    - ZIP file: extracts and scans inside
    - Excel file: uses directly
    Returns: (excel_file_paths_list, temp_dirs_to_cleanup_list)
    """
    excel_files = []
    temp_dirs = []
    
    print(f"\n🔍 Scanning {len(paths)} path(s)...")
    
    for path in paths:
        p = Path(path)
        
        if not p.exists():
            print(f"\n   ❌ Path not found: {path}")
            continue
        
        if p.is_file():
            # Single file
            if p.suffix.lower() == '.zip':
                extracted, temp_dir = extract_zip(str(p))
                excel_files.extend(extracted)
                if temp_dir:
                    temp_dirs.append(temp_dir)
                    
            elif p.suffix.lower() in ['.xlsx', '.xls']:
                excel_files.append(str(p.resolve()))
                
        elif p.is_dir():
            # Folder: find Excel files directly inside
            direct_excel = []
            for ext in ['*.xlsx', '*.xls', '*.XLSX', '*.XLS']:
                direct_excel.extend(p.rglob(ext))
            
            for f in direct_excel:
                excel_files.append(str(f.resolve()))
            
            print(f"\n   📁 Folder: {p.name}")
            print(f"      ✅ Direct Excel files: {len(direct_excel)}")
            
            # Folder: also find ZIP files inside and extract them
            zip_files = []
            for ext in ['*.zip', '*.ZIP']:
                zip_files.extend(p.rglob(ext))
            
            print(f"      📦 ZIP files found: {len(zip_files)}")
            
            for zip_file in zip_files:
                extracted, temp_dir = extract_zip(str(zip_file))
                excel_files.extend(extracted)
                if temp_dir:
                    temp_dirs.append(temp_dir)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_files = []
    for f in excel_files:
        if f not in seen:
            seen.add(f)
            unique_files.append(f)
    
    print(f"\n📋 Total Excel files to process: {len(unique_files)}")
    return unique_files, temp_dirs

# ========== COLUMN MAPPER ==========
def find_column(columns_list, keywords):
    for col in columns_list:
        col_lower = str(col).lower().strip()
        for keyword in keywords:
            if keyword in col_lower:
                return col
    return None

def get_column_mapping(columns_list):
    mapping = {}
    used_cols = set()
    for master_col, keywords in COLUMN_KEYWORDS.items():
        matched_col = find_column(columns_list, keywords)
        if matched_col and matched_col not in used_cols:
            mapping[master_col] = matched_col
            used_cols.add(matched_col)
    return mapping

# ========== STEP 1: COPY ALL ROWS ==========
def extract_from_file(file_path):
    print(f"\n📄 Processing: {os.path.basename(file_path)}")
    
    try:
        engine = 'xlrd' if str(file_path).lower().endswith('.xls') else 'openpyxl'
        df = pd.read_excel(file_path, engine=engine)
    except Exception as e:
        # .xls file may actually be XLSX format internally — retry with openpyxl
        if str(file_path).lower().endswith('.xls'):
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
            except Exception as e2:
                print(f"   ❌ Error: {e2}")
                return None
        else:
            print(f"   ❌ Error: {e}")
            return None
    
    if df.empty:
        print(f"   ⚠️ Empty file")
        return None
    
    columns = list(df.columns)
    mapping = get_column_mapping(columns)
    print(f"   Mapping: {mapping}")
    
    # 🔴 CHANGED: PAN No → PAN No.
    if 'PAN No.' not in mapping:
        print(f"   ❌ PAN not found in: {columns}")
        return None
    
    master_df = pd.DataFrame()
    # 🔴 CHANGED: All PAN No references now use PAN No.
    master_df['PAN No.'] = df[mapping['PAN No.']].astype(str).str.strip().str.upper()
    master_df['Remark'] = df[mapping['Remark']].astype(str).str.strip() if 'Remark' in mapping else ''
    master_df['Notice'] = df[mapping['Notice']].astype(str).str.strip() if 'Notice' in mapping else ''
    
    print(f"   ✅ Copied {len(master_df)} rows")
    return master_df

# ========== STEP 2: MERGE ==========
def merge_all_files(file_paths):
    all_dfs = []
    for file_path in file_paths:
        df = extract_from_file(file_path)
        if df is not None and not df.empty:
            all_dfs.append(df)
    
    if not all_dfs:
        print("\n❌ No valid data!")
        return pd.DataFrame(columns=MASTER_COLUMNS)
    
    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\n📊 Step 1: All copied. Rows: {len(combined)}")
    return combined

# ========== STEPS 2-4: CLEAN ==========
def clean_master_table(df):
    if df.empty:
        return df
    
    df = df.copy()
    
    print("\n" + "=" * 60)
    print("STEP 2: Remove duplicates (PAN) - First pass")
    print("=" * 60)
    before = len(df)
    # 🔴 CHANGED: PAN No → PAN No.
    df = df.drop_duplicates(subset=['PAN No.'], keep='first')
    print(f"🗑️  Removed {before - len(df)}. Rows: {len(df)}")
    
    print("\n" + "=" * 60)
    print("STEP 3: Remove spaces from PAN No.")
    print("=" * 60)
    # 🔴 CHANGED: PAN No → PAN No.
    df['PAN No.'] = df['PAN No.'].str.replace(' ', '', regex=False)
    print(f"✂️  Done")
    
    print("\n" + "=" * 60)
    print("STEP 4: Remove duplicates (PAN No.) - Second pass")
    print("=" * 60)
    before = len(df)
    # 🔴 CHANGED: PAN No → PAN No.
    df = df.drop_duplicates(subset=['PAN No.'], keep='first')
    print(f"🗑️  Removed {before - len(df)}. Rows: {len(df)}")
    
    # 🔴 CHANGED: PAN No → PAN No.
    df = df[df['PAN No.'].notna() & (df['PAN No.'] != '') & (df['PAN No.'] != 'NAN')]
    print(f"✅ Final: {len(df)}")
    
    return df.reset_index(drop=True)

# ========== CLEANUP TEMP DIRS ==========
def cleanup_temp_dirs(temp_dirs):
    print("\n🧹 Cleaning up temporary extracted files...")
    for temp_dir in temp_dirs:
        try:
            shutil.rmtree(temp_dir)
            print(f"   ✅ Removed: {temp_dir}")
        except Exception as e:
            print(f"   ⚠️ Could not remove {temp_dir}: {e}")





# ========== NEW: API FUNCTIONS ==========
def upload_to_api(file_path, api_url=None, api_key=None):
    """Upload file to IndiaBonds API."""
    url = api_url or API_URL
    key = api_key or API_KEY
    headers = {"X-Api-Key": key}
    params = {"Comment": COMMENT}

    with open(file_path, "rb") as f:
        files = {"file": ("SEBIDebarred_Format.xlsx", f,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = requests.post(url, headers=headers, params=params, files=files, timeout=60)

    return response

def handle_api_response(response):
    """Parse API response. Returns dict with status and error info."""
    content_type = response.headers.get("Content-Type", "")
    
    if "application/json" not in content_type:
        return {"status": "unknown", "message": f"Unexpected content type: {content_type}", 
                "needs_filter": False, "error_file": None, "errors": []}
    
    data = response.json()
    
    # Success - no errors
    if data.get("result") == "OK":
        return {"status": "OK", "message": "200 OK - All PANs valid", 
                "needs_filter": False, "error_file": None, "errors": []}
    
    # Has errors
    if data.get("result") == "ProcessedWithErrors":
        errors = data.get("errors", [])
        error_file_b64 = data.get("errorFile")
        error_bytes = base64.b64decode(error_file_b64) if error_file_b64 else None
        
        return {
            "status": "ProcessedWithErrors",
            "message": f"Found {len(errors)} invalid PAN(s)",
            "needs_filter": True,
            "error_file": error_bytes,  # Raw bytes for download
            "errors": errors
        }
    
    return {"status": "unknown", "message": "Unknown response format", 
            "needs_filter": False, "error_file": None, "errors": []}

def filter_and_reupload(error_bytes, api_url=None, api_key=None):
    """Filter invalid PANs from error file and re-upload clean version."""
    url = api_url or API_URL
    key = api_key or API_KEY
    df = pd.read_excel(BytesIO(error_bytes))
    print(f"\n📊 Error file: {len(df)} rows")

    df_clean = df[df['Error Description'].isna() |
                  (df['Error Description'].astype(str).str.strip() != 'InValid PAN Value')].copy()
    df_clean = df_clean.drop(columns=['Error Description'])
    df_clean = df_clean[['PAN No.', 'Remark', 'Notice']]

    removed = len(df) - len(df_clean)
    print(f"🔍 Filtered: {len(df)} → {len(df_clean)} rows (removed {removed} invalid)")

    clean_buffer = BytesIO()
    df_clean.to_excel(clean_buffer, index=False, engine='openpyxl')
    clean_buffer.seek(0)

    print("🚀 Re-uploading clean file...")
    headers = {"X-Api-Key": key}
    params = {"Comment": COMMENT}
    files = {"file": ("SEBIDebarred_Format.xlsx", clean_buffer,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    response = requests.post(url, headers=headers, params=params, files=files, timeout=60)

    result = handle_api_response(response)
    return {
        "final_status": result["status"],
        "final_message": result["message"],
        "removed_count": removed,
        "clean_df": df_clean,
        "clean_bytes": clean_buffer.getvalue()
    }

# ========== YOUR EXISTING run_pipeline() WITH API ADDED AT END ==========
def run_pipeline(paths, output_path='SEBIDebarred_Format.xlsx', api_url=None, api_key=None):
    print("=" * 60)
    print("EXCEL MASTER TABLE PROCESSOR")
    print("=" * 60)
    
    excel_files, temp_dirs = get_excel_files(paths)
    
    if not excel_files:
        print("\n❌ No Excel files!")
        cleanup_temp_dirs(temp_dirs)
        return None
    
    print(f"\n📋 Files:")
    for i, f in enumerate(excel_files, 1):
        print(f"   {i}. {os.path.basename(str(f))}")
    
    print("\n" + "=" * 60)
    print("STEP 1: Copy all rows")
    print("=" * 60)
    combined = merge_all_files(excel_files)
    
    if combined.empty:
        print("\n❌ Nothing to save!")
        cleanup_temp_dirs(temp_dirs)
        return None
    
    cleaned = clean_master_table(combined)
    
    print("\n" + "=" * 60)
    print("STEP 5: Save")
    print("=" * 60)
    
    # Save dataframe first
    cleaned.to_excel(output_path, index=False)

    # Load workbook
    wb = load_workbook(output_path)
    ws = wb.active

    # Get table range
    last_row = ws.max_row
    last_col = ws.max_column
    table_range = f"A1:{chr(64 + last_col)}{last_row}"

    # Create table
    tab = Table(displayName="MasterTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    wb.save(output_path)

    print("✅ Excel table styled successfully")
    print(f"💾 Saved: {os.path.abspath(output_path)}")
    
    # ========== NEW: STEP 6 - UPLOAD TO API ==========
    print("\n" + "=" * 60)
    print("STEP 6: Upload to API")
    print("=" * 60)
    
    api_response = upload_to_api(output_path, api_url=api_url, api_key=api_key)
    api_result = handle_api_response(api_response)
    
    print(f"📡 API Status: {api_result['status']}")
    print(f"📡 Message: {api_result['message']}")
    cleanup_temp_dirs(temp_dirs)
    # Return dict with everything Streamlit needs
    return {
        "master_df": cleaned,
        "output_path": output_path,
        "api_status": api_result["status"],
        "api_message": api_result["message"],
        "needs_filter": api_result["needs_filter"],
        "error_file": api_result["error_file"],      # Bytes for download button
        "errors": api_result["errors"],              # List of error dicts
        "can_proceed": api_result["needs_filter"]      # True if user needs to decide
    }

